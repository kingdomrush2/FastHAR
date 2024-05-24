import random

import numpy as np
import torch
from openpyxl import Workbook
from openpyxl.styles import Font
import os
import itertools
from scipy.stats import special_ortho_group


def merge_dataset(data, label):
    index = np.zeros(data.shape[0], dtype=bool)
    label_new = []
    for i in range(label.shape[0]):
        temp_label = np.unique(label[i])
        if temp_label.size == 1:
            index[i] = True
            label_new.append(label[i, 0])
    return data[index], np.array(label_new)


def select_data(data, labels):
    user_label_index, position_label_index = 1, 2
    data = data[labels[:, 0, position_label_index] == 2, ...]
    labels = labels[labels[:, 0, position_label_index] == 2, ...]
    return data, labels


def split_data_label(data, labels, train_rate, valid_rate):
    # data, labels = select_data(data, labels)

    label_index = 0
    arr = np.arange(data.shape[0])
    np.random.shuffle(arr)
    data = data[arr]
    labels = labels[arr]
    train_num = int(data.shape[0] * train_rate)
    vali_num = int(data.shape[0] * valid_rate)
    data_train = data[:train_num, ...]
    data_vali = data[train_num:train_num + vali_num, ...]
    data_test = data[train_num + vali_num:, ...]
    t = np.min(labels[:, :, label_index])
    label_train = labels[:train_num, ..., label_index] - t
    label_vali = labels[train_num:train_num + vali_num, ..., label_index] - t
    label_test = labels[train_num + vali_num:, ..., label_index] - t

    data_train, label_train = merge_dataset(data_train, label_train)
    data_test, label_test = merge_dataset(data_test, label_test)
    data_vali, label_vali = merge_dataset(data_vali, label_vali)
    print('Train Size: %d, Vali Size: %d, Test Size: %d' % (
    label_train.shape[0], label_vali.shape[0], label_test.shape[0]))
    return data_train, label_train, data_vali, label_vali, data_test, label_test


def split_data_label_tv(data, labels, train_rate=0.5714):
    label_index = 0
    arr = np.arange(data.shape[0])
    np.random.shuffle(arr)
    data = data[arr]
    labels = labels[arr]
    train_num = int(data.shape[0] * train_rate)

    data_train = data[:train_num, ...]
    data_vali = data[train_num:, ...]
    t = np.min(labels[:, :, label_index])
    label_train = labels[:train_num, ..., label_index] - t
    label_vali = labels[train_num:, ..., label_index] - t

    data_train, label_train = merge_dataset(data_train, label_train)
    data_vali, label_vali = merge_dataset(data_vali, label_vali)
    print('Train Size: %d, Vali Size: %d' % (label_train.shape[0], label_vali.shape[0]))
    return data_train, label_train, data_vali, label_vali


def down_sampling(data, original_sampling_rate, aim_sampling_rate):
    step = int(original_sampling_rate / aim_sampling_rate)
    result_data = data[:, ::step, :]
    return result_data


class WriteExcel:
    def __init__(self, excel_name):
        self.excel_name = excel_name
        self.letter_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']
        self.dataset_list = ['uci', 'shoaib', 'motion', 'hhar']
        self.init_excel()

    def init_excel(self):
        self.wb = Workbook()
        self.ws = self.wb.active

        self.ws.merge_cells('A1:A2')
        for i, column in enumerate(range(1, 13, 3)):
            self.ws.merge_cells(self.letter_list[column] + '1:' + self.letter_list[column + 2] + '1')
            self.ws[self.letter_list[column] + '1'] = self.dataset_list[i]
            self.ws[self.letter_list[column] + '2'] = 'cm'
            self.ws[self.letter_list[column + 1] + '2'] = 'acc'
            self.ws[self.letter_list[column + 2] + '2'] = 'f1'
        self.ws.merge_cells('N1:O1')
        self.ws['N1'] = 'avg'
        self.ws['N2'] = 'acc'
        self.ws['O2'] = 'f1'
        self.ws['A7'] = 'avg_all'
        for i in range(len(self.dataset_list)):
            self.ws['A' + str(i + 3)] = self.dataset_list[i]

    def write_excel(self, source_dataset, target_dataset, cm, acc, f1):
        row_index = self.dataset_list.index(source_dataset) + 3
        column_index = self.dataset_list.index(target_dataset) * 3 + 1
        font = Font(size=7)
        self.ws[self.letter_list[column_index] + str(row_index)] = self.deal_cm(cm.replace(' ', ','))
        self.ws[self.letter_list[column_index] + str(row_index)].font = font
        self.ws[self.letter_list[column_index + 1] + str(row_index)] = str(round(acc * 100, 2)) + "%"
        self.ws[self.letter_list[column_index + 2] + str(row_index)] = str(round(f1 * 100, 2)) + "%"

    def write_excel_avg(self, dataset, acc, f1):
        if dataset in self.dataset_list:
            row_index = self.dataset_list.index(dataset) + 3
        else:
            row_index = 7  # write avg_all
        acc_column, f1_column = 'N', 'O'
        self.ws[acc_column + str(row_index)] = str(round(acc * 100, 2)) + "%"
        self.ws[f1_column + str(row_index)] = str(round(f1 * 100, 2)) + "%"

    def deal_cm(self, cm):
        stage = 1
        result = ""
        for i in range(len(cm)):
            if cm[i] != ',':
                result += cm[i]
                if stage == 1:
                    stage = 0
            else:
                if stage == 1:
                    result += " "
                elif cm[i - 1] == '\n' or cm[i - 1] == '[':
                    pass
                else:
                    result += ","
                    stage = 1
        return result

    def save_excel(self, save_path):
        self.wb.save(os.path.join(save_path, self.excel_name))


def augument_dataset(data, label, method='channel_aug'):
    # data(sample num, sequence_len, feature), label(sample num, sequence_len, feature)
    print(f'begin data augmentation: method is {method}')
    data_res = np.empty((0, data.shape[1], data.shape[2]))
    label_res = np.empty((0, label.shape[1], label.shape[2]))
    aug_num = 5  # aug_num: the number of sample of one sample will augument

    if method == '':
        data_res = np.concatenate([data_res, data], axis=0)
        label_res = np.concatenate([label_res, label], axis=0)
        return data_res, label_res

    elif method == 'rotation_random':
        axis_num = 3

        # each sample rotation matrix is same
        for i in range(aug_num):
            axis = np.random.uniform(low=-1, high=1, size=axis_num)
            angle = np.random.uniform(low=-np.pi, high=np.pi)
            rotation_mat = special_ortho_group.rvs(3)
            data_temp = data.reshape(-1, 6)
            aug_temp_acc = np.matmul(data_temp[:, :3], rotation_mat)
            aug_temp_gyr = np.matmul(data_temp[:, 3:], rotation_mat)
            aug_temp = np.concatenate([aug_temp_acc, aug_temp_gyr], axis=1)
            aug_temp = aug_temp.reshape(-1, 120, 6)
            data_res = np.concatenate([data_res, aug_temp], axis=0)
            label_res = np.concatenate([label_res, label], axis=0)
        data_res = np.concatenate([data_res, data], axis=0)
        label_res = np.concatenate([label_res, label], axis=0)

    else:
        print('method not exist')
        return None, None

    print(f'data augumentation end')
    return data_res, label_res


def make_one_shot_dataset(data, label, activity_label_index=0):
    _, index = np.unique(label, return_index=True)
    res_data, res_label = data[index], label[index]
    return res_data, res_label


def change_windowsize(data, label, window_size):
    # data (num, windowsize, feature) label (num, window_size, label_num)

    if data.shape[1] == window_size:
        return data, label

    print(f'change window_size from {data.shape[1]} to {window_size}')
    label_num = label.shape[2]
    label_unique_list = []
    for i in range(label_num):
        unique_value, count = np.unique(label[:, 0, i], return_counts=True)
        label_unique_list.append(unique_value.tolist())

    res_data = np.empty(shape=(0, window_size, data.shape[2]))
    res_label = np.empty(shape=(0, window_size, label_num))
    label_combines = list(itertools.product(*label_unique_list))
    for label_combine in label_combines:
        label_combine = np.array(label_combine)
        data_temp = data[(label[:, 0, :] == label_combine).all(axis=1)]

        data_temp = data_temp.reshape(-1, data.shape[2])
        redundancy = data_temp.shape[0] % window_size
        data_temp = data_temp[:-redundancy, :]
        data_temp = data_temp.reshape(-1, window_size, data.shape[2])

        label_temp = np.zeros(shape=(data_temp.shape[0], data_temp.shape[1], label_num))
        for i in range(len(label_combine)):
            label_temp[:, :, i] = label_combine[i]
        res_data = np.concatenate([res_data, data_temp], axis=0)
        res_label = np.concatenate([res_label, label_temp], axis=0)

    print(f'data shape after change window_size{res_data.shape}')
    print(f'label shape after change window_size{res_label.shape}')

    return res_data, res_label
